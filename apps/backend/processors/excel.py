"""Procesador de archivos Excel (.xlsx, .xls, .csv)."""
import hashlib
import logging
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd
import xlrd
from core.errors import AppError, ErrorCode
from core.models import ExtractedContent, FileType, TableData
from core.session import Session
from lxml import etree

logger = logging.getLogger(__name__)

MAX_ROWS_FULL = 500       # Hojas con más filas se samplearán
MAX_SHEETS = 10           # Máximo de hojas a procesar
MAX_COLS = 50             # Máximo de columnas por hoja


def process_excel(file_path: Path, session: Session) -> ExtractedContent:
    suffix = file_path.suffix.lower()
    try:
        if suffix == ".csv":
            return _process_csv(file_path, session)
        elif suffix == ".xls":
            return _process_xls(file_path, session)
        else:
            return _process_xlsx(file_path, session)
    except AppError:
        raise
    except Exception as e:
        logger.error("Error procesando Excel: %s", type(e).__name__)
        raise AppError(ErrorCode.CORRUPT_FILE, str(e))


# ─── XLSX ─────────────────────────────────────────────────────────────────────

def _process_xlsx(file_path: Path, session: Session) -> ExtractedContent:
    warnings: list[str] = []
    tables: list[TableData] = []
    metadata: dict = {}
    text_parts: list[str] = []

    # Abrir con data_only=True para obtener valores cacheados
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    except Exception:
        raise AppError(ErrorCode.CORRUPT_FILE)

    sheet_names = wb.sheetnames[:MAX_SHEETS]
    metadata["hojas"] = sheet_names
    metadata["total_hojas"] = len(wb.sheetnames)

    if len(wb.sheetnames) > MAX_SHEETS:
        warnings.append(f"El archivo tiene {len(wb.sheetnames)} hojas. Se procesaron las primeras {MAX_SHEETS}.")

    # Metadata de conexiones externas y tablas dinámicas desde XML interno
    _extract_xml_metadata(file_path, metadata, warnings)

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            warnings.append(f"Hoja '{sheet_name}' está oculta — incluida de todas formas.")

        df = _worksheet_to_dataframe(ws, warnings, sheet_name)
        if df is None or df.empty:
            continue

        table_data, sheet_warnings = _dataframe_to_table(df, sheet_name)
        warnings.extend(sheet_warnings)
        tables.append(table_data)

        # Resumen estadístico de columnas numéricas
        stats = _describe_dataframe(df, sheet_name)
        if stats:
            text_parts.append(stats)

    wb.close()

    text_content = "\n\n".join(text_parts) if text_parts else "No se encontraron datos en el archivo."
    return ExtractedContent(
        source_file_hash=hashlib.sha256(file_path.name.encode()).hexdigest()[:16],
        file_type=FileType.EXCEL,
        text_content=text_content,
        tables=tables,
        metadata=metadata,
        extraction_warnings=warnings,
    )


def _worksheet_to_dataframe(ws, warnings: list, sheet_name: str) -> pd.DataFrame | None:
    """Convierte una hoja de openpyxl a DataFrame, detectando el header real."""
    try:
        data = list(ws.values)
        if not data:
            return None

        # Buscar la fila del header (primera fila con ≥3 celdas no vacías con strings)
        header_row = 0
        for i, row in enumerate(data[:20]):
            non_empty = [c for c in row if c is not None]
            if len(non_empty) >= 3:
                string_ratio = sum(1 for c in non_empty if isinstance(c, str)) / len(non_empty)
                if string_ratio > 0.5:
                    header_row = i
                    break

        headers = [str(c) if c is not None else f"Col_{j}" for j, c in enumerate(data[header_row])]
        headers = headers[:MAX_COLS]
        rows = data[header_row + 1:]
        # Truncar/rellenar filas para que coincidan exactamente con MAX_COLS
        trimmed_rows = [list(r)[:MAX_COLS] + [''] * max(0, len(headers) - len(list(r))) for r in rows]

        df = pd.DataFrame(trimmed_rows, columns=headers)
        df = df.dropna(how="all")

        # Samplear si es muy grande
        if len(df) > MAX_ROWS_FULL:
            warnings.append(
                f"Hoja '{sheet_name}': {len(df):,} filas — se analizó una muestra representativa de {MAX_ROWS_FULL}."
            )
            df = _smart_sample(df, MAX_ROWS_FULL)

        return df
    except Exception as e:
        logger.warning("Error convirtiendo hoja '%s': %s", sheet_name, type(e).__name__)
        return None


def _smart_sample(df: pd.DataFrame, target: int) -> pd.DataFrame:
    """Muestra inteligente: primeras + últimas + aleatorias del medio."""
    n = len(df)
    head = df.head(100)
    tail = df.tail(50)
    middle_n = max(0, target - 150)
    middle = df.iloc[100:-50].sample(min(middle_n, max(0, n - 150)), random_state=42) if n > 150 else pd.DataFrame()
    return pd.concat([head, middle, tail]).drop_duplicates()


def _dataframe_to_table(df: pd.DataFrame, sheet_name: str) -> tuple[TableData, list[str]]:
    warnings = []
    headers = list(df.columns.astype(str))
    rows = []
    for _, row in df.iterrows():
        rows.append([str(v) if v is not None and not (isinstance(v, float) and pd.isna(v)) else "" for v in row])
    return TableData(headers=headers, rows=rows, sheet_name=sheet_name), warnings


def _describe_dataframe(df: pd.DataFrame, sheet_name: str) -> str:
    """Genera un resumen estadístico legible de columnas numéricas."""
    numeric_cols = df.select_dtypes(include="number")
    if numeric_cols.empty:
        return ""
    desc = numeric_cols.describe().round(2).to_string()
    return f"Estadísticas de hoja '{sheet_name}':\n{desc}"


def _extract_xml_metadata(file_path: Path, metadata: dict, warnings: list) -> None:
    """Lee el XML interno del .xlsx para extraer tablas dinámicas y conexiones."""
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            names = zf.namelist()

            # Tablas dinámicas
            pivot_count = sum(1 for n in names if "pivotTable" in n)
            if pivot_count:
                metadata["tablas_dinamicas"] = pivot_count
                warnings.append(
                    f"Se detectaron {pivot_count} tablas dinámicas. Se analizaron los datos fuente directamente."
                )

            # Conexiones externas
            if "xl/connections.xml" in names:
                with zf.open("xl/connections.xml") as f:
                    tree = etree.parse(f)
                    conns = tree.findall(".//{*}connection")
                    if conns:
                        conn_names = [c.get("name", "sin nombre") for c in conns[:5]]
                        metadata["conexiones_externas"] = conn_names
                        warnings.append(
                            f"El archivo tiene {len(conns)} conexión(es) a fuentes externas (no resueltas): "
                            + ", ".join(conn_names[:3])
                        )

            # Rangos con nombre
            if "xl/workbook.xml" in names:
                with zf.open("xl/workbook.xml") as f:
                    tree = etree.parse(f)
                    defined_names = tree.findall(".//{*}definedName")
                    if defined_names:
                        metadata["rangos_nombrados"] = [n.get("name") for n in defined_names[:10]]

    except Exception as e:
        logger.debug("No se pudo leer XML interno: %s", type(e).__name__)


# ─── CSV ──────────────────────────────────────────────────────────────────────

def _process_csv(file_path: Path, session: Session) -> ExtractedContent:
    warnings: list[str] = []
    try:
        # Detectar encoding
        for enc in ("utf-8", "latin-1", "cp1252"):
            try:
                df = pd.read_csv(file_path, encoding=enc, nrows=10)
                df = pd.read_csv(file_path, encoding=enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise AppError(ErrorCode.CORRUPT_FILE)
    except AppError:
        raise
    except Exception:
        raise AppError(ErrorCode.CORRUPT_FILE)

    if len(df) > MAX_ROWS_FULL:
        warnings.append(f"CSV con {len(df):,} filas — se analizó una muestra de {MAX_ROWS_FULL}.")
        df = _smart_sample(df, MAX_ROWS_FULL)

    table, _ = _dataframe_to_table(df, "datos")
    stats = _describe_dataframe(df, "datos")

    return ExtractedContent(
        source_file_hash=hashlib.sha256(file_path.name.encode()).hexdigest()[:16],
        file_type=FileType.EXCEL,
        text_content=stats or f"CSV con {len(df)} filas y {len(df.columns)} columnas.",
        tables=[table],
        metadata={"columnas": list(df.columns), "filas": len(df)},
        extraction_warnings=warnings,
    )


# ─── XLS (formato legacy) ─────────────────────────────────────────────────────

def _process_xls(file_path: Path, session: Session) -> ExtractedContent:
    warnings: list[str] = []
    tables: list[TableData] = []
    text_parts: list[str] = []

    try:
        wb = xlrd.open_workbook(str(file_path))
    except Exception:
        raise AppError(ErrorCode.CORRUPT_FILE)

    sheet_names = wb.sheet_names()[:MAX_SHEETS]
    warnings.append("Formato .xls (Excel 97-2003). Fórmulas no disponibles — se usan valores guardados.")

    for name in sheet_names:
        ws = wb.sheet_by_name(name)
        if ws.nrows < 2:
            continue

        headers = [str(ws.cell_value(0, c)) or f"Col_{c}" for c in range(min(ws.ncols, MAX_COLS))]
        rows = []
        for r in range(1, min(ws.nrows, MAX_ROWS_FULL + 1)):
            rows.append([str(ws.cell_value(r, c)) for c in range(min(ws.ncols, MAX_COLS))])

        tables.append(TableData(headers=headers, rows=rows, sheet_name=name))
        if ws.nrows > MAX_ROWS_FULL:
            warnings.append(f"Hoja '{name}': {ws.nrows} filas — se procesaron las primeras {MAX_ROWS_FULL}.")

        df = pd.DataFrame(rows, columns=headers)
        stats = _describe_dataframe(df, name)
        if stats:
            text_parts.append(stats)

    return ExtractedContent(
        source_file_hash=hashlib.sha256(file_path.name.encode()).hexdigest()[:16],
        file_type=FileType.EXCEL,
        text_content="\n\n".join(text_parts) or "Archivo Excel legacy sin datos detectables.",
        tables=tables,
        metadata={"hojas": sheet_names, "formato": "xls"},
        extraction_warnings=warnings,
    )
