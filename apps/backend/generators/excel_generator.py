"""Generador de archivos Excel a partir del análisis de IA."""
import logging
import re
from pathlib import Path

import openpyxl
from core.models import AnalysisResult
from core.session import Session
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Paleta corporativa
COLOR_HEADER_BG = "1F497D"
COLOR_HEADER_FG = "FFFFFF"
COLOR_ROW_ALT   = "EBF3FB"
COLOR_BORDER    = "BDD7EE"


def generate_excel(result: AnalysisResult, session: Session) -> Path:
    """
    Genera un archivo Excel con tablas y gráficos a partir del texto de análisis.
    El texto de la IA debe contener bloques ---TABLE--- con formato CSV-like.
    Si no hay tablas estructuradas, genera una hoja con el texto como sumario.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Quitar hoja vacía por defecto

    tables = _parse_tables(result.analysis_text)

    if tables:
        for i, tbl in enumerate(tables):
            sheet_name = tbl.get("titulo", f"Tabla {i+1}")[:31]  # máx 31 chars en Excel
            ws = wb.create_sheet(title=sheet_name)
            _write_table(ws, tbl)
            if tbl.get("grafico") and len(tbl["rows"]) > 1:
                _add_bar_chart(ws, tbl, sheet_name)
    else:
        # Fallback: hoja de resumen con el texto del análisis
        ws = wb.create_sheet(title="Análisis")
        _write_summary(ws, result.analysis_text)

    # Hoja de metadata
    ws_meta = wb.create_sheet(title="Info")
    ws_meta["A1"] = "Generado por Excel Analyzer"
    ws_meta["A1"].font = Font(bold=True, color=COLOR_HEADER_BG)
    ws_meta["A2"] = f"Sesión: {result.session_id[:8]}..."
    ws_meta["A3"] = "Advertencias:"
    for i, w in enumerate(result.warnings, start=4):
        ws_meta[f"A{i}"] = f"• {w}"

    out_path = session.temp_dir / "datos_analisis.xlsx"
    wb.save(str(out_path))
    logger.info("Excel generado: %s", out_path)
    return out_path


# ─────────────────────────────────────────────────────────────────
# Parser
# ─────────────────────────────────────────────────────────────────

def _parse_tables(text: str) -> list[dict]:
    """
    Parsea bloques ---TABLE--- del texto de la IA.
    Formato esperado:
        TITULO: Nombre de la tabla
        GRAFICO: si|no
        ENCABEZADOS: Col1, Col2, Col3
        Col1, Col2, Col3
        Val1, Val2, Val3
    """
    blocks = re.split(r"---TABLE---", text, flags=re.IGNORECASE)
    tables = []
    for raw in blocks:
        raw = raw.strip()
        if not raw:
            continue
        tbl: dict = {"titulo": "Tabla", "grafico": False, "headers": [], "rows": []}
        lines = raw.splitlines()
        data_started = False
        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue
            upper = stripped.upper()
            if upper.startswith("TITULO:"):
                tbl["titulo"] = stripped.split(":", 1)[1].strip()
            elif upper.startswith("GRAFICO:"):
                tbl["grafico"] = stripped.split(":", 1)[1].strip().lower() in ("si", "sí", "yes", "true", "1")
            elif upper.startswith("ENCABEZADOS:"):
                tbl["headers"] = [h.strip() for h in stripped.split(":", 1)[1].split(",")]
                data_started = True
            elif data_started or ("," in stripped and not upper.startswith(("TIPO:", "TITULO:", "GRAFICO:", "ENCABEZADOS:"))):
                cells = [c.strip() for c in stripped.split(",")]
                if cells:
                    tbl["rows"].append(cells)

        if tbl["headers"] and tbl["rows"]:
            tables.append(tbl)

    return tables


# ─────────────────────────────────────────────────────────────────
# Escritura de tablas
# ─────────────────────────────────────────────────────────────────

def _write_table(ws, tbl: dict) -> None:
    thin = Side(style="thin", color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = tbl["headers"]
    # Fila de encabezado
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color=COLOR_HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    # Filas de datos
    for row_idx, row in enumerate(tbl["rows"], start=2):
        fill = PatternFill("solid", fgColor=COLOR_ROW_ALT) if row_idx % 2 == 0 else None
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col, value=_parse_value(val))
            if fill:
                cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)

    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    ws.freeze_panes = "A2"


def _add_bar_chart(ws, tbl: dict, title: str) -> None:
    """Agrega un gráfico de barras debajo de la tabla."""
    n_rows = len(tbl["rows"]) + 1  # +1 por encabezado
    n_cols = len(tbl["headers"])

    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = 10
    chart.height = 12
    chart.width = 20

    # Usar columnas numéricas (cols 2..n) como series
    data = Reference(ws, min_col=2, max_col=n_cols, min_row=1, max_row=n_rows)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    anchor_row = n_rows + 3
    ws.add_chart(chart, f"A{anchor_row}")


def _write_summary(ws, text: str) -> None:
    """Escribe el análisis como texto en una hoja cuando no hay tablas estructuradas."""
    ws.column_dimensions["A"].width = 100
    ws["A1"] = "Resumen del Análisis"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_HEADER_BG)

    row = 3
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            row += 1
            continue
        cell = ws.cell(row=row, column=1, value=stripped)
        if stripped.startswith("## "):
            cell.value = stripped[3:]
            cell.font = Font(bold=True, size=12, color=COLOR_HEADER_BG)
        elif stripped.startswith("# "):
            cell.value = stripped[2:]
            cell.font = Font(bold=True, size=13, color=COLOR_HEADER_BG)
        elif stripped.startswith(("- ", "* ")):
            cell.value = "  • " + stripped[2:]
        cell.alignment = Alignment(wrap_text=True)
        row += 1


def _parse_value(val: str):
    """Intenta convertir a número para que Excel lo trate como dato numérico."""
    v = val.replace(",", "").replace("%", "").strip()
    try:
        return int(v)
    except ValueError:
        pass
    try:
        return float(v)
    except ValueError:
        return val
